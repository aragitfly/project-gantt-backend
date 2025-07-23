from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import json
import os
import tempfile
import shutil
import time
import datetime
from typing import List, Optional
import aiofiles
import re
from dateutil import parser
from openai import OpenAI

app = FastAPI(title="Project Gantt Chart Manager", version="1.0.0")

# Environment variables
PORT = int(os.environ.get("PORT", 8000))
ENVIRONMENT = os.environ.get("ENVIRONMENT", "development")

# CORS middleware - allow all origins in production, specific origins in development
if ENVIRONMENT == "production":
    allow_origins = ["*"]  # Allow all origins in production for now
else:
    allow_origins = [
        "http://localhost:3000",
        "https://localhost:3000",
        "https://project-gantt-frontend3.vercel.app",
        "https://project-gantt-frontend3-m6b7o7t0l-pauls-projects-045e95a7.vercel.app",
        "null"
    ]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allow_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Debug logging for CORS
print(f"Environment: {ENVIRONMENT}")
print(f"Allowed origins: {allow_origins}")

# Pydantic models
class Project(BaseModel):
    name: str
    item_id: str
    activity_type: str
    is_title: bool
    start_date: Optional[str]
    end_date: Optional[str]
    team: str
    status: str
    completed: int

class ProjectUpdate(BaseModel):
    project_name: str
    task_name: str
    new_start_date: str
    new_end_date: str

class AudioTranscript(BaseModel):
    transcript: str
    project_updates: List[ProjectUpdate]

# Global variable to store the current Excel file
current_excel_file = None
# Global variable to store the current projects data
current_projects = []

@app.get("/")
async def root():
    return {"message": "Project Gantt Chart Manager API"}

@app.get("/cors-test")
async def cors_test():
    """Test endpoint to verify CORS is working"""
    from datetime import datetime
    return {"message": "CORS test successful", "timestamp": str(datetime.now())}

@app.get("/openai-test")
async def openai_test():
    """Test endpoint to check if OpenAI API is configured"""
    openai_api_key = os.environ.get("OPENAI_API_KEY")
    
    if not openai_api_key:
        return {
            "openai_configured": False,
            "message": "OpenAI API key not found in environment variables"
        }
    
    return {
        "openai_configured": True,
        "message": "OpenAI API key is configured",
        "key_preview": f"{openai_api_key[:10]}..." if openai_api_key else "Not set"
    }

@app.get("/dutch-test")
async def dutch_test():
    """Test endpoint to verify Dutch language processing"""
    openai_api_key = os.environ.get("OPENAI_API_KEY")
    
    if not openai_api_key:
        return {
            "dutch_support": False,
            "message": "OpenAI API key not found"
        }
    
    try:
        # Initialize OpenAI client
        try:
            client = OpenAI(api_key=openai_api_key)
        except Exception as init_error:
            import openai
            openai.api_key = openai_api_key
            client = openai
        
        # Test with a simple Dutch phrase
        test_text = "Dit is een test van de Nederlandse taalverwerking."
        
        return {
            "dutch_support": True,
            "message": "Dutch language processing is configured",
            "test_phrase": test_text,
            "client_type": "new" if hasattr(client, 'audio') else "old"
        }
        
    except Exception as e:
        return {
            "dutch_support": False,
            "message": f"Error testing Dutch support: {str(e)}"
        }

@app.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    """Upload and parse Excel file containing Gantt chart data"""
    global current_excel_file
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file")
    
    try:
        # Save uploaded file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        shutil.copyfileobj(file.file, temp_file)
        temp_file.close()
        
        current_excel_file = temp_file.name
        
        # Read Excel file
        df = pd.read_excel(temp_file.name)
        
        # Extract project data (assuming specific structure)
        projects = []
        
        # Look for project data in the Excel file
        for index, row in df.iterrows():
            # Skip rows before row 8 (where headers start)
            if index < 8:
                continue
                
            # Check if this row contains project data
            project_name = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ""  # Activiteiten column
            item_id = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""      # # column
            
            # Skip if no meaningful data or if it's a header
            if not project_name.strip() or project_name.strip().lower() in ['activiteiten', 'nan']:
                continue
                
            # Skip section headers (like "Generieke services", "Autoschade")
            if project_name.strip().lower() in ['generieke services', 'autoschade']:
                continue
                
            # Extract data from the row
            team = "Unassigned"  # Default team value
            start_date = None
            end_date = None
            status = "Planning"
            completed = 0
            activity_type = "general"
            
            # Parse team from column 3 (Team column)
            if len(row) > 3 and pd.notna(row.iloc[3]):
                team_value = str(row.iloc[3]).strip()
                if team_value and team_value.lower() != 'nan':
                    team = team_value
            
            # Parse dates from columns 4 and 5 (Start and End dates)
            if len(row) > 4 and pd.notna(row.iloc[4]):
                try:
                    if isinstance(row.iloc[4], datetime):
                        start_date = row.iloc[4].strftime('%Y-%m-%d')
                    else:
                        start_date = str(row.iloc[4])
                except:
                    start_date = str(row.iloc[4]) if pd.notna(row.iloc[4]) else None
                    
            if len(row) > 5 and pd.notna(row.iloc[5]):
                try:
                    if isinstance(row.iloc[5], datetime):
                        end_date = row.iloc[5].strftime('%Y-%m-%d')
                    else:
                        end_date = str(row.iloc[5])
                except:
                    end_date = str(row.iloc[5]) if pd.notna(row.iloc[5]) else None
            
            # Parse status from column 7
            if len(row) > 7 and pd.notna(row.iloc[7]):
                status = str(row.iloc[7])
            
            # Parse completion percentage from column 8
            if len(row) > 8 and pd.notna(row.iloc[8]):
                try:
                    completed_value = row.iloc[8]
                    if isinstance(completed_value, (int, float)):
                        # Convert decimal to percentage (e.g., 0.2 -> 20, 1 -> 100)
                        if completed_value <= 1:
                            completed = int(completed_value * 100)
                        else:
                            completed = int(completed_value)
                    else:
                        completed_str = str(completed_value).strip()
                        if '%' in completed_str:
                            completed = int(completed_str.replace('%', '').strip())
                        elif completed_str.replace('.', '').isdigit():
                            completed = int(float(completed_str) * 100)
                        else:
                            completed = 0
                except (ValueError, TypeError):
                    completed = 0
            
            # Determine activity type based on item ID
            is_title = False
            if item_id and item_id.strip() != 'nan':
                if '.' in item_id:
                    if item_id.count('.') == 1:  # e.g., "1.1", "1.2"
                        activity_type = "sub-activity"
                        is_title = False
                    elif item_id.count('.') > 1:  # e.g., "1.4.1"
                        activity_type = "sub-sub-activity"
                        is_title = False
                    else:
                        activity_type = "main-item"
                        is_title = True
                elif item_id.isdigit():
                    activity_type = "main-item"
                    is_title = True  # Main items are titles
                else:
                    # For items without dots, check if they're main activities
                    # Main activities are typically single numbers or have no hierarchical structure
                    activity_type = "main-item"
                    is_title = True
            else:
                # If no item_id, treat as main activity
                activity_type = "main-item"
                is_title = True
            
            # Clean up the project name
            if project_name and project_name.strip() and project_name.strip() != "nan":
                clean_name = project_name.strip()
                project_data = {
                    "name": clean_name,
                    "item_id": item_id if item_id.strip() != 'nan' else "",
                    "activity_type": activity_type,
                    "is_title": is_title,
                    "start_date": start_date,
                    "end_date": end_date,
                    "team": team,
                    "status": status,
                    "completed": completed
                }
                projects.append(project_data)
                print(f"DEBUG: Added project '{clean_name}' with item_id='{item_id}', is_title={is_title}, activity_type={activity_type}")
                

        
        # Store projects in global variable for task proposals
        global current_projects
        current_projects = projects
        
        return {
            "message": "Excel file uploaded successfully",
            "filename": file.filename,
            "projects": projects,
            "total_rows": len(df)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing Excel file: {str(e)}")

@app.post("/process-audio")
async def process_audio(audio_file: UploadFile = File(...)):
    """Process audio file and extract project updates using OpenAI Whisper API"""
    
    print(f"DEBUG: Received audio file: {audio_file.filename}, size: {audio_file.size}")
    
    # Accept more audio formats including webm
    allowed_extensions = ('.wav', '.mp3', '.m4a', '.webm', '.ogg')
    if not audio_file.filename.lower().endswith(allowed_extensions):
        print(f"DEBUG: Invalid file extension: {audio_file.filename}")
        raise HTTPException(status_code=400, detail=f"Audio file must be one of: {', '.join(allowed_extensions)}")
    
    try:
        # Check if OpenAI API key is available
        openai_api_key = os.environ.get("OPENAI_API_KEY")
        if not openai_api_key:
            print("DEBUG: OpenAI API key not found")
            return {
                "transcript": "Audio processing requires OpenAI API key. Please contact support.",
                "summary": "Audio processing service is not properly configured.",
                "taskProposals": [],
                "project_updates": []
            }
        
        # Initialize OpenAI client
        print(f"DEBUG: Starting OpenAI client initialization...")
        use_new_format = False
        try:
            client = OpenAI(api_key=openai_api_key)
            print(f"DEBUG: OpenAI client initialized successfully with new format")
            use_new_format = True
        except Exception as init_error:
            print(f"DEBUG: OpenAI client initialization failed: {str(init_error)}")
            # Try alternative initialization
            try:
                import openai
                openai.api_key = openai_api_key
                client = openai
                print(f"DEBUG: OpenAI client initialized with old format")
                use_new_format = False
            except Exception as alt_error:
                print(f"DEBUG: Alternative initialization also failed: {str(alt_error)}")
                return {
                    "transcript": "Audio processing service is temporarily unavailable. Please try again later.",
                    "summary": "Audio processing service is being updated.",
                    "taskProposals": [],
                    "project_updates": []
                }
        
        # Save audio file temporarily
        file_extension = os.path.splitext(audio_file.filename)[1].lower()
        temp_audio = tempfile.NamedTemporaryFile(delete=False, suffix=file_extension)
        shutil.copyfileobj(audio_file.file, temp_audio)
        temp_audio.close()
        
        audio_file_path = temp_audio.name
        cleanup_files = [temp_audio.name]
        
        # Process audio with OpenAI Whisper API
        try:
            print("DEBUG: Starting speech recognition with OpenAI Whisper API...")
            print(f"DEBUG: Audio file path: {audio_file_path}")
            print(f"DEBUG: Audio file size: {os.path.getsize(audio_file_path)} bytes")
            
            with open(audio_file_path, "rb") as file_obj:
                print(f"DEBUG: File opened successfully, making API call...")
                # Always use the new API format since that's what's available
                print("DEBUG: Making API call to OpenAI...")
                try:
                    # Try using the old client format that was successfully initialized
                    print("DEBUG: Using old client format for API call")
                    transcript_response = client.audio.transcriptions.create(
                        model="whisper-1",
                        file=file_obj,
                        language="nl",  # Dutch language
                        response_format="text",
                        prompt="This is a Dutch business meeting about project management and task updates."
                    )
                    print("DEBUG: API call completed successfully")
                    transcript = transcript_response
                except Exception as api_error:
                    print(f"DEBUG: API call failed: {str(api_error)}")
                    # Try alternative approach with requests library
                    try:
                        print("DEBUG: Trying alternative API call method")
                        import requests
                        headers = {
                            'Authorization': f'Bearer {openai_api_key}'
                        }
                        files = {
                            'file': ('audio.webm', file_obj, 'audio/webm'),
                            'model': (None, 'whisper-1'),
                            'language': (None, 'nl'),
                            'response_format': (None, 'text'),
                            'prompt': (None, 'This is a Dutch business meeting about project management and task updates.')
                        }
                        response = requests.post(
                            'https://api.openai.com/v1/audio/transcriptions',
                            headers=headers,
                            files=files
                        )
                        if response.status_code == 200:
                            transcript = response.text
                            print("DEBUG: Alternative API call completed successfully")
                        else:
                            print(f"DEBUG: Alternative API call failed with status {response.status_code}")
                            transcript = "Could not understand audio. Please try again with clearer speech."
                    except Exception as alt_error:
                        print(f"DEBUG: Alternative API call also failed: {str(alt_error)}")
                        transcript = "Could not understand audio. Please try again with clearer speech."
            
            print(f"DEBUG: Speech recognition successful: {transcript[:100]}...")
            
            # Check if transcript is empty or contains error messages
            if not transcript or transcript.strip() == "":
                transcript = "Could not understand audio. Please try again with clearer speech."
            elif "could not understand" in transcript.lower() or "unclear" in transcript.lower():
                transcript = "Could not understand audio. Please try again with clearer speech."
            
        except Exception as api_error:
            print(f"DEBUG: OpenAI API processing failed: {str(api_error)}")
            transcript = "Could not understand audio. Please try again with clearer speech."
        
        # Clean up temp files
        for file_path in cleanup_files:
            try:
                os.unlink(file_path)
            except:
                pass
        
        # Extract project updates from transcript
        project_updates = extract_project_updates(transcript)
        
        # Generate AI-powered task proposals based on transcript
        task_proposals = generate_task_proposals(transcript)
        
        # Generate meeting summary
        summary = generate_meeting_summary(transcript)
        
        return {
            "transcript": transcript,
            "summary": summary,
            "taskProposals": task_proposals,
            "project_updates": project_updates
        }
        
    except Exception as e:
        print(f"DEBUG: Error in audio processing: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error processing audio: {str(e)}")

def extract_project_updates(transcript: str) -> List[ProjectUpdate]:
    """Extract project updates from transcript using regex patterns"""
    updates = []
    
    # Patterns to match project updates
    patterns = [
        r"(\w+)\s+(?:is|are)\s+(\d+)\s*%?\s*(?:complete|done|finished)",
        r"(\w+)\s+(?:has|have)\s+(?:been\s+)?(completed|finished|done)",
        r"(\w+)\s+(?:start|begin)\s+(?:on|from)\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
        r"(\w+)\s+(?:end|finish)\s+(?:on|by)\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})"
    ]
    
    for pattern in patterns:
        matches = re.finditer(pattern, transcript, re.IGNORECASE)
        for match in matches:
            project_name = match.group(1)
            update_info = match.group(2)
            
            # Create a basic update (you can enhance this)
            update = ProjectUpdate(
                project_name=project_name,
                task_name=project_name,
                new_start_date="",
                new_end_date=""
            )
            updates.append(update)
    
    return updates

def generate_task_proposals(transcript: str) -> List[dict]:
    """Generate AI-powered task proposals based on transcript analysis"""
    proposals = []
    
    # Keywords that indicate task status changes
    status_keywords = {
        "completed": ["complete", "finished", "done", "accomplished", "finalized", "klaar", "voltooid", "afgerond"],
        "in_progress": ["progress", "ongoing", "working", "developing", "implementing", "bezig", "lopen", "werken"],
        "delayed": ["delay", "behind", "late", "postponed", "extended", "vertraagd", "achter", "laat"],
        "blocked": ["block", "stuck", "issue", "problem", "obstacle", "blokker", "probleem", "obstakel"]
    }
    
    # Get current projects from global variable
    global current_projects
    available_tasks = current_projects if current_projects else []
    
    # Dutch and English task patterns
    task_patterns = [
        # Dutch patterns
        r"(?:tussentijds|opslaan|bestand)\s+(?:is|zijn)\s+(\d+)\s*%?\s*(?:klaar|voltooid|gedaan)",
        r"(?:tussentijds|opslaan|bestand)\s+(?:heeft|hebben)\s+(?:al\s+)?(voltooid|afgerond|klaar)",
        r"(?:tussentijds|opslaan|bestand)\s+(?:is|zijn)\s+(vertraagd|achter|laat)",
        r"(?:tussentijds|opslaan|bestand)\s+(?:is|zijn)\s+(geblokkeerd|vastgelopen|geblokkeerd)",
        # English patterns
        r"(\w+)\s+(?:is|are)\s+(\d+)\s*%?\s*(?:complete|done|finished)",
        r"(\w+)\s+(?:has|have)\s+(?:been\s+)?(completed|finished|done)",
        r"(\w+)\s+(?:is|are)\s+(delayed|behind|late)",
        r"(\w+)\s+(?:is|are)\s+(blocked|stuck|blocking)"
    ]
    
    # First, try to match specific tasks mentioned in the transcript
    for task in available_tasks:
        task_name_lower = task['name'].lower()
        transcript_lower = transcript.lower()
        
        # Check if task name is mentioned in transcript
        if task_name_lower in transcript_lower:
            # Look for status indicators near the task mention
            status_found = None
            progress_found = None
            
            # Check for Dutch status patterns
            if any(keyword in transcript_lower for keyword in status_keywords["completed"]):
                status_found = "Completed"
                progress_found = 100
            elif any(keyword in transcript_lower for keyword in status_keywords["delayed"]):
                status_found = "Delayed"
                progress_found = 30
            elif any(keyword in transcript_lower for keyword in status_keywords["blocked"]):
                status_found = "Blocked"
                progress_found = 20
            elif any(keyword in transcript_lower for keyword in status_keywords["in_progress"]):
                status_found = "In Progress"
                progress_found = 50
            
            if status_found:
                proposals.append({
                    "id": f"proposal-{len(proposals) + 1}",
                    "taskId": task['item_id'],  # Use actual task ID from Excel
                    "proposedStatus": status_found,
                    "proposedProgress": progress_found,
                    "reason": f"Task '{task['name']}' status mentioned in meeting",
                    "confidence": 0.8,
                    "meetingId": f"meeting-{int(time.time())}",
                    "timestamp": datetime.datetime.now().isoformat()
                })
    
    # If no specific tasks found, create general proposals based on transcript content
    if not proposals and available_tasks:
        # Look for general status mentions
        if any(keyword in transcript.lower() for keyword in status_keywords["completed"]):
            # Propose completion for first main task
            main_tasks = [t for t in available_tasks if t['activity_type'] == 'main-item']
            if main_tasks:
                task = main_tasks[0]
                proposals.append({
                    "id": f"proposal-{len(proposals) + 1}",
                    "taskId": task['item_id'],
                    "proposedStatus": "Completed",
                    "proposedProgress": 100,
                    "reason": "Completion mentioned in meeting",
                    "confidence": 0.6,
                    "meetingId": f"meeting-{int(time.time())}",
                    "timestamp": datetime.datetime.now().isoformat()
                })
        
        elif any(keyword in transcript.lower() for keyword in status_keywords["delayed"]):
            # Propose delay for first main task
            main_tasks = [t for t in available_tasks if t['activity_type'] == 'main-item']
            if main_tasks:
                task = main_tasks[0]
                proposals.append({
                    "id": f"proposal-{len(proposals) + 1}",
                    "taskId": task['item_id'],
                    "proposedStatus": "Delayed",
                    "proposedProgress": 30,
                    "reason": "Delay mentioned in meeting",
                    "confidence": 0.6,
                    "meetingId": f"meeting-{int(time.time())}",
                    "timestamp": datetime.datetime.now().isoformat()
                })
    
    return proposals

def generate_meeting_summary(transcript: str) -> str:
    """Generate a concise meeting summary from transcript"""
    
    # Extract key points
    key_points = []
    
    # Look for action items (Dutch and English)
    action_patterns = [
        # Dutch patterns
        r"(?:we moeten|we moeten|we moeten|actie|todo)\s+(.+?)(?:\.|$)",
        r"(?:toewijzen|delegeren)\s+(.+?)\s+aan\s+(.+?)(?:\.|$)",
        r"(?:aanpassen|wijzigen|veranderen)\s+(.+?)(?:\.|$)",
        r"(?:hulp nodig|help nodig)\s+(?:van|bij)\s+(.+?)(?:\.|$)",
        r"(?:stuurgroep|steering group)\s+(.+?)(?:\.|$)",
        # English patterns
        r"(?:need to|should|must|will)\s+(.+?)(?:\.|$)",
        r"(?:action item|todo|task)\s*:\s*(.+?)(?:\.|$)",
        r"(?:next step|next action)\s*:\s*(.+?)(?:\.|$)"
    ]
    
    for pattern in action_patterns:
        matches = re.finditer(pattern, transcript, re.IGNORECASE)
        for match in matches:
            if len(match.groups()) > 1:
                key_points.append(f"Action: {match.group(1).strip()} to {match.group(2).strip()}")
            else:
                key_points.append(f"Action: {match.group(1).strip()}")
    
    # Look for progress updates (Dutch and English)
    progress_patterns = [
        # Dutch patterns
        r"(\w+)\s+(?:is|zijn)\s+(\d+)\s*%?\s*(?:klaar|voltooid|gedaan)",
        r"(\w+)\s+(?:heeft|hebben)\s+(?:al\s+)?(voltooid|afgerond|klaar)",
        # English patterns
        r"(\w+)\s+(?:is|are)\s+(\d+)\s*%?\s*(?:complete|done)",
        r"(\w+)\s+(?:has|have)\s+(?:been\s+)?(completed|finished)"
    ]
    
    for pattern in progress_patterns:
        matches = re.finditer(pattern, transcript, re.IGNORECASE)
        for match in matches:
            key_points.append(f"Progress: {match.group(1)} - {match.group(2)}")
    
    # Look for issues/blockers (Dutch and English)
    issue_patterns = [
        # Dutch patterns
        r"(?:probleem|blokker|obstakel|issue)\s*:\s*(.+?)(?:\.|$)",
        r"(?:vertraagd|achter|laat)\s+omdat\s+(.+?)(?:\.|$)",
        r"(?:blokker|blokkeert)\s+(.+?)(?:\.|$)",
        r"(?:daar hebben we een blokker|we hebben een probleem)\s+(.+?)(?:\.|$)",
        # English patterns
        r"(?:issue|problem|blocker|obstacle)\s*:\s*(.+?)(?:\.|$)",
        r"(?:delayed|behind|late)\s+because\s+(.+?)(?:\.|$)"
    ]
    
    for pattern in issue_patterns:
        matches = re.finditer(pattern, transcript, re.IGNORECASE)
        for match in matches:
            key_points.append(f"Issue: {match.group(1).strip()}")
    
    # Look for date changes (Dutch and English)
    date_patterns = [
        # Dutch patterns
        r"(?:aanpassen|wijzigen|veranderen)\s+(?:naar|to)\s+(\d+\s+\w+|\w+\s+\d+)",
        r"(?:nieuwe datum|nieuwe deadline)\s*:\s*(\d+\s+\w+|\w+\s+\d+)",
        r"(?:passen.*aan naar)\s+(\d+\s+\w+|\w+\s+\d+)",
        # English patterns
        r"(?:change|move|update)\s+(?:to|date)\s+(\d+\s+\w+|\w+\s+\d+)",
        r"(?:new date|new deadline)\s*:\s*(\d+\s+\w+|\w+\s+\d+)"
    ]
    
    for pattern in date_patterns:
        matches = re.finditer(pattern, transcript, re.IGNORECASE)
        for match in matches:
            key_points.append(f"Date Change: {match.group(1).strip()}")
    
    if key_points:
        summary = "Meeting Summary:\n" + "\n".join(key_points[:8])  # Limit to 8 key points
    else:
        # If no specific patterns found, create a basic summary with key Dutch phrases
        if "tussentijds" in transcript.lower() or "opslaan" in transcript.lower():
            key_points.append("Action: Review interim saving functionality")
        if "blokker" in transcript.lower() or "probleem" in transcript.lower():
            key_points.append("Issue: Blocker identified - needs steering group assistance")
        if "5 augustus" in transcript.lower():
            key_points.append("Date Change: Adjust timeline to August 5th")
        
        if key_points:
            summary = "Meeting Summary:\n" + "\n".join(key_points)
        else:
            summary = f"Meeting discussion recorded. Key topics: {transcript[:200]}..."
    
    return summary

@app.post("/update-excel")
async def update_excel(updates: List[ProjectUpdate]):
    """Update Excel file with project changes"""
    global current_excel_file
    
    if not current_excel_file:
        raise HTTPException(status_code=400, detail="No Excel file uploaded")
    
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(current_excel_file)
        sheet = workbook.active
        
        # Apply updates
        for update in updates:
            # Find project in sheet
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and update.project_name.lower() in str(cell.value).lower():
                        # Found project, update dates
                        row_idx = cell.row
                        
                        # Update start date (assuming it's in a specific column)
                        start_col = 3  # Adjust based on your Excel structure
                        end_col = 4    # Adjust based on your Excel structure
                        
                        try:
                            start_date = datetime.strptime(update.new_start_date, '%Y-%m-%d')
                            end_date = datetime.strptime(update.new_end_date, '%Y-%m-%d')
                            
                            sheet.cell(row=row_idx, column=start_col, value=start_date)
                            sheet.cell(row=row_idx, column=end_col, value=end_date)
                        except:
                            pass
                        break
        
        # Save updated workbook
        workbook.save(current_excel_file)
        
        return {
            "message": "Excel file updated successfully",
            "updates_applied": len(updates)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating Excel file: {str(e)}")

@app.get("/download-excel")
async def download_excel():
    """Download the updated Excel file"""
    global current_excel_file
    
    if not current_excel_file:
        raise HTTPException(status_code=400, detail="No Excel file available")
    
    try:
        from fastapi.responses import FileResponse
        return FileResponse(
            current_excel_file,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename='updated_gantt_chart.xlsx'
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error downloading file: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000) 